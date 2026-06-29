# -*- coding: utf-8 -*-
"""
导出当前数据库中的基础数据表。

数据源：PostgreSQL (grid_data + weather_obs)

运行：
    python export_base_table.py
    → data/v11_15min_base.xlsx
"""
import sys, io, numpy as np, pandas as pd
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

PROJECT = Path(__file__).resolve().parent
OUTPUT = PROJECT / "data"
OUTPUT.mkdir(parents=True, exist_ok=True)

from shared.weather_config import CLUSTERS, CLUSTER_LABELS, VAR_PREFIXES, NODE_NAMES

# ── 集群/变量标签 ──
VAR_LABELS = {
    "temp": "气温(℃)", "rh": "湿度(%)", "precip": "降水(mm/h)",
    "cloud": "云量(0-1)", "wind": "100m风速(m/s)", "rad": "辐射(W/m²)",
}

# ── 流域→集群映射（降水）──
BASIN_CLUSTERS = {
    "金沙江下游": ["panzhihua"],
    "雅砻江流域": ["liangshan", "panzhihua"],
    "大渡河流域": ["yaan"],
    "岷江流域":   ["yaan"],
}

# ── 气温区域→集群映射 ──
TEMP_CLUSTERS = {
    "成都平原":   "chengdu",
    "川东城市群": "dazhou",
    "川南城市群": "yibin",
}


def _load_from_db():
    """从 PostgreSQL 读取 grid_data + weather_obs，合并为 DataFrame。"""
    from config import settings
    from sqlalchemy import create_engine
    import json

    url = settings.database_url_sync
    engine = create_engine(url, echo=False)
    print(f"数据库: {url.split('@')[1] if '@' in url else url}")

    # Grid data
    df_grid = pd.read_sql(
        "SELECT datetime, price, load, solar, wind, hydro, "
        "renewable_total, bidspace, reserve, nonmarket, tieline, load_tie, day_type "
        "FROM grid_data ORDER BY datetime",
        engine,
    )
    df_grid["datetime"] = pd.to_datetime(df_grid["datetime"])
    print(f"  grid_data: {len(df_grid)} 行")

    # Weather data
    df_w = pd.read_sql(
        "SELECT datetime, variables FROM weather_obs ORDER BY datetime",
        engine,
    )
    df_w["datetime"] = pd.to_datetime(df_w["datetime"])
    print(f"  weather_obs: {len(df_w)} 行")

    # Expand JSONB
    weather_expanded = []
    for _, row in df_w.iterrows():
        d = {"datetime": row["datetime"]}
        vars_dict = row["variables"]
        if isinstance(vars_dict, str):
            vars_dict = json.loads(vars_dict)
        d.update(vars_dict)
        weather_expanded.append(d)

    df_weather = pd.DataFrame(weather_expanded)
    n_weather_cols = len(df_weather.columns) - 1  # minus datetime
    print(f"  weather keys: {n_weather_cols}")

    # Merge + ffill
    df = df_grid.merge(df_weather, on="datetime", how="left")
    df = df.set_index("datetime").sort_index()

    weather_cols_in = [c for c in df_weather.columns if c != "datetime" and c in df.columns]
    df[weather_cols_in] = df[weather_cols_in].ffill()

    engine.dispose()
    return df


def export():
    # ═══════════════════════════════════════════════════════════════
    # 1. 从 PostgreSQL 加载数据
    # ═══════════════════════════════════════════════════════════════
    print("=" * 60)
    print("  基础数据表导出 — v11 (Open-Meteo ECMWF / 19 子节点集群平均)")
    print("=" * 60)

    df = _load_from_db()
    source = f"PostgreSQL ({len(df)} 行)"

    dt = df.index
    n = len(df)
    print(f"  合并: {n} 行 × {len(df.columns)} 列")
    print(f"  日期范围: {dt.min()} → {dt.max()}")

    # ═══════════════════════════════════════════════════════════════
    # 2. 工具函数
    # ═══════════════════════════════════════════════════════════════
    hour = dt.hour.values
    minute = dt.minute.values
    period = hour * 4 + minute // 15

    def _cmean(prefix, cluster):
        """集群内等权平均。"""
        total, cnt = np.zeros(n), 0
        for node in CLUSTERS[cluster]:
            col = f"{prefix}_{node}"
            if col in df.columns:
                total += np.nan_to_num(df[col].values, nan=0.0)
                cnt += 1
        return total / cnt if cnt > 0 else np.full(n, np.nan)

    # ═══════════════════════════════════════════════════════════════
    # 3. 构建输出
    # ═══════════════════════════════════════════════════════════════
    out = pd.DataFrame(index=range(n))

    # ── A. 基础电网 (13 列) ──
    out["日期"] = dt.strftime("%Y-%m-%d")
    out["时段"] = [f"{h:02d}:{m:02d}-{h:02d}:{(m+15)%60:02d}"
                    for h, m in zip(hour, minute)]
    out["运行阶段"] = "正式运行"
    out["日期类型"] = df["day_type"].values if "day_type" in df.columns else "工作日"
    out["新能源总出力(MW)"] = df["renewable_total"].values if "renewable_total" in df.columns else np.nan
    out["水电(MW)"] = df["hydro"].values
    out["风电(MW)"] = df["wind"].values
    out["光伏(MW)"] = df["solar"].values
    out["省内负荷(MW)"] = df["load"].values
    out["出清价格(元/MWh)"] = df["price"].values
    out["数据来源"] = "Open-Meteo ECMWF v3 (19子节点集群平均)"
    out["离群标记"] = np.nan
    out["datetime_bj"] = dt.strftime("%Y-%m-%d %H:%M:%S")

    # ── B. 时间特征 (9 列) ──
    m = dt.month.values
    out["枯水期"] = ((m >= 1) & (m <= 4)).astype(int)
    out["汛期"]   = (m >= 5).astype(int)
    out["主汛期"] = (m >= 7).astype(int)
    sf = np.datetime64("2026-02-17")
    days_sf = (dt.values - sf).astype("timedelta64[D]").astype(float)
    out["节前1天"]  = (days_sf == -1).astype(int)
    out["节后1天"]  = (days_sf == 1).astype(int)
    out["早高峰"]   = ((period >= 20) & (period < 36)).astype(int)
    out["晚高峰"]   = ((period >= 68) & (period < 88)).astype(int)
    out["午间谷段"] = ((period >= 48) & (period < 68)).astype(int)

    # ── C. 天气-集群平均 (7 集群 × 6 变量 = 42 列) ──
    for cluster, label in CLUSTER_LABELS.items():
        for prefix in VAR_PREFIXES:
            col_name = f"{label}_{VAR_LABELS[prefix]}"
            out[col_name] = _cmean(prefix, cluster)
            if prefix == "cloud":
                # 云量: 原始值 0-100 → 0-1
                out[col_name] = out[col_name] / 100.0

    # ── D. 天气-衍生 (16 列) ──
    # D1. 24h 降水 × 4 流域 (hourly-aware rolling sum)
    def _roll24h_basin(clusters):
        total = np.zeros(n)
        for cl in clusters:
            total += _cmean("precip", cl)
        # hourly-aware: downsample → roll → upsample
        arr_h = total[::4]
        hs = pd.Series(arr_h).rolling(24, min_periods=1).sum().values
        return np.repeat(hs, 4)[:n]

    for basin, clusters in BASIN_CLUSTERS.items():
        out[f"{basin}_24h降水(mm)"] = _roll24h_basin(clusters)

    # D2. 24h 气温变化 × 3 区域
    for region, cluster in TEMP_CLUSTERS.items():
        t = _cmean("temp", cluster)
        out[f"{region}_气温24h变(℃)"] = t - np.roll(t, 96)

    # D3. HDD/CDD × 3 区域
    for region, cluster in TEMP_CLUSTERS.items():
        t = _cmean("temp", cluster)
        out[f"{region}_HDD"] = np.maximum(18.0 - t, 0)
        out[f"{region}_CDD"] = np.maximum(t - 18.0, 0)

    out["气象数据"] = 1

    # ── E. 电网市场 (5 列) ──
    out["竞价空间(MW)"]  = df["bidspace"].values if "bidspace" in df.columns else np.nan
    out["系统备用(MW)"]  = df["reserve"].values if "reserve" in df.columns else np.nan
    out["非市场机组(MW)"] = df["nonmarket"].values if "nonmarket" in df.columns else np.nan
    out["联络线(MW)"]    = df["tieline"].values if "tieline" in df.columns else np.nan
    out["负荷联络线(MW)"] = df["load_tie"].values if "load_tie" in df.columns else np.nan

    # ═══════════════════════════════════════════════════════════════
    # 4. 列分组（保存前确定，供后续格式化和列清单使用）
    # ═══════════════════════════════════════════════════════════════
    def _assign_group(col_name):
        if col_name in ["日期", "时段", "运行阶段", "日期类型", "数据来源", "离群标记", "datetime_bj"]:
            return "A.基础电网"
        if any(k in col_name for k in ["新能源", "水电(MW)", "风电(MW)", "光伏(MW)", "省内负荷", "出清价格"]):
            return "A.基础电网"
        if any(k in col_name for k in ["枯水", "汛期", "节前", "节后", "高峰", "谷段"]):
            return "B.时间特征"
        if any(k in col_name for k in ["24h降水", "气温24h变", "HDD", "CDD", "气象数据"]):
            return "D.天气-衍生"
        if any(k in col_name for k in ["_气温", "_湿度", "_降水", "_云量", "风速", "_辐射"]):
            return "C.天气-集群"
        if any(k in col_name for k in ["竞价", "备用", "非市场", "联络线"]):
            return "E.电网市场"
        return "?"

    groups = [_assign_group(c) for c in out.columns]
    ca = sum(1 for g in groups if g == "A.基础电网")
    cb = sum(1 for g in groups if g == "B.时间特征")
    cc = sum(1 for g in groups if g == "C.天气-集群")
    cd = sum(1 for g in groups if g == "D.天气-衍生")
    ce = sum(1 for g in groups if g == "E.电网市场")

    # ═══════════════════════════════════════════════════════════════
    # 5. 写入 Excel（含格式）
    # ═══════════════════════════════════════════════════════════════
    fp = OUTPUT / "v11_15min_base.xlsx"
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers,
    )
    from openpyxl.utils import get_column_letter

    # ── 配色方案 ──
    GROUP_FILLS = {
        "A.基础电网": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),   # 浅蓝
        "B.时间特征": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),   # 浅绿
        "C.天气-集群": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),   # 浅橙
        "D.天气-衍生": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),   # 浅黄
        "E.电网市场": PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),   # 浅紫
    }
    HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
    DATA_FONT = Font(name="微软雅黑", size=9)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    DATA_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # ── 按列名推断数字格式 ──
    def _col_fmt(col_name):
        """根据列名推断合适的数字格式。"""
        if any(k in col_name for k in ["价格"]):
            return "#,##0.00"
        if any(k in col_name for k in ["出力", "水电", "风电", "光伏", "负荷", "竞价", "备用", "非市场", "联络线"]):
            return "#,##0"
        if any(k in col_name for k in ["气温", "温度"]):
            return "0.0"
        if any(k in col_name for k in ["湿度", "云量"]):
            return "0.0"
        if any(k in col_name for k in ["降水"]):
            return "0.00"
        if any(k in col_name for k in ["风速", "辐射"]):
            return "0.0"
        if any(k in col_name for k in ["HDD", "CDD"]):
            return "0.0"
        if any(k in col_name for k in ["24h变"]):
            return "0.0"
        if any(k in col_name for k in ["枯水", "汛期", "节前", "节后", "高峰", "谷段", "气象数据"]):
            return "0"
        return None

    with pd.ExcelWriter(fp, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="基础数据", index=False, startrow=0)

        # 元信息 sheet
        meta = pd.DataFrame({
            "项目": ["数据源", "天气API", "节点数", "集群数", "变量数", "JSONB key数",
                    "导出时间", "行数", "列数", "日期范围", "特征版本"],
            "值": [source,
                  "Open-Meteo ECMWF IFS (archive-api + forecast)",
                  f"{len(NODE_NAMES)} 子节点",
                  f"{len(CLUSTERS)} 区域集群",
                  f"{len(VAR_PREFIXES)} (temp/rh/precip/cloud/wind/rad)",
                  f"{len(NODE_NAMES) * len(VAR_PREFIXES)}",
                  pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
                  f"{n}",
                  f"{len(out.columns)}",
                  f"{dt.min()} → {dt.max()}",
                  "v11"],
        })
        meta.to_excel(writer, sheet_name="元信息", index=False)

        # 节点列表 sheet
        nodes_df = pd.DataFrame([
            {"集群": CLUSTER_LABELS.get(cl, cl), "子节点": node}
            for cl, nodes in CLUSTERS.items()
            for node in nodes
        ])
        nodes_df.to_excel(writer, sheet_name="节点列表", index=False)

        # 列清单 sheet
        col_list = pd.DataFrame({
            "序号": range(1, len(out.columns) + 1),
            "列名": out.columns.tolist(),
            "分组": groups,
        })
        col_list.to_excel(writer, sheet_name="列清单", index=False)

    # ═══════════════════════════════════════════════════════════════
    # 6. 格式美化（openpyxl 二次打开）
    # ═══════════════════════════════════════════════════════════════
    from openpyxl import load_workbook

    wb = load_workbook(fp)

    # ── 6a. 基础数据 sheet ──
    ws = wb["基础数据"]
    n_cols = len(out.columns)
    n_rows = n

    # 冻结：行1（表头）+ 列A-B（日期、时段）
    ws.freeze_panes = "C2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"

    # ── 列宽（自适应） ──
    # 先为每列计算合理宽度
    col_widths = {}
    for col_idx, col_name in enumerate(out.columns, 1):
        # 表头宽度
        header_w = 0
        for ch in str(col_name):
            header_w += 2.2 if ord(ch) > 127 else 1.2
        # 数据抽样宽度（取前 100 行 + 每隔 1000 行取一行）
        data_max = 0
        sample_indices = list(range(min(100, n_rows)))
        if n_rows > 100:
            sample_indices += list(range(100, n_rows, max(1, n_rows // 50)))[:50]
        for i in sample_indices:
            val = out.iloc[i, col_idx - 1]
            if pd.isna(val):
                continue
            s = str(val)
            w = 0
            for ch in s:
                w += 2.2 if ord(ch) > 127 else 1.1
            data_max = max(data_max, w)
        col_widths[col_idx] = min(max(header_w, data_max) + 4, 28)

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 表头格式 ──
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        grp = groups[col_idx - 1]
        fill = GROUP_FILLS.get(grp, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # ── 数据区域格式 ──
    # 数字格式: 前 3000 行（~31 天 15-min 数据）+ 每 500 行一个抽样点
    NUM_FMT_ROWS = min(3000, n_rows)
    fmt_indices = set(range(2, NUM_FMT_ROWS + 2))
    for r in range(NUM_FMT_ROWS + 2, n_rows + 2, 500):
        fmt_indices.add(r)

    for col_idx in range(1, n_cols + 1):
        fmt = _col_fmt(out.columns[col_idx - 1])
        if fmt:
            for r in fmt_indices:
                ws.cell(row=r, column=col_idx).number_format = fmt

    # 可视格式: 前 500 行 + 稀疏抽样（字体/边框/对齐）
    VISIBLE = min(500, n_rows)
    visible_rows = set(range(2, VISIBLE + 2))
    step = max(1, n_rows // 60)
    for r in range(VISIBLE + 2, n_rows + 2, step):
        visible_rows.add(r)

    for row_idx in visible_rows:
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_idx <= 2:
                cell.alignment = DATA_ALIGN_CENTER
            else:
                cell.alignment = DATA_ALIGN_RIGHT

    # ── 行高 ──
    ws.row_dimensions[1].height = 28  # 表头行高

    # ── 6b. 元信息 sheet ──
    ws_meta = wb["元信息"]
    ws_meta.column_dimensions["A"].width = 16
    ws_meta.column_dimensions["B"].width = 52
    for row in ws_meta.iter_rows(min_row=1, max_row=ws_meta.max_row,
                                  min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = Font(name="微软雅黑", bold=True, size=10)
                cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # ── 6c. 节点列表 sheet ──
    ws_nodes = wb["节点列表"]
    ws_nodes.column_dimensions["A"].width = 14
    ws_nodes.column_dimensions["B"].width = 14
    for row in ws_nodes.iter_rows(min_row=1, max_row=ws_nodes.max_row,
                                   min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = Font(name="微软雅黑", bold=True, size=10)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # ── 6d. 列清单 sheet ──
    ws_cols = wb["列清单"]
    ws_cols.column_dimensions["A"].width = 8
    ws_cols.column_dimensions["B"].width = 28
    ws_cols.column_dimensions["C"].width = 16
    for row in ws_cols.iter_rows(min_row=1, max_row=ws_cols.max_row,
                                  min_col=1, max_col=3):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = Font(name="微软雅黑", bold=True, size=10)
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            # 给列清单的分组列也着色
            if cell.row > 1 and cell.column == 3:
                grp = cell.value
                if grp in GROUP_FILLS:
                    cell.fill = GROUP_FILLS[grp]

    wb.save(fp)
    wb.close()

    # ═══════════════════════════════════════════════════════════════
    # 7. 控制台总结
    # ═══════════════════════════════════════════════════════════════
    print(f"""
  导出: {fp}
  {n} 行 × {len(out.columns)} 列 (4 sheets)
  格式: 分组着色表头 | 冻结C2 | 自动筛选 | 自适应列宽 | 数字格式

╔══════════════════════════════════════════════════════════╗
║        v11 基础表列结构 (共 {len(out.columns)} 列)              ║
╠══════════════════════════════════════════════════════════╣
║  A. 基础电网   {ca:3d} 列   蓝  日期/时段/出力/价格          ║
║  B. 时间特征   {cb:3d} 列   绿  季节/节日/峰谷时段            ║
║  C. 天气-集群  {cc:3d} 列   橙  7集群×6变量 集群内等权平均    ║
║  D. 天气-衍生  {cd:3d} 列   黄  24h降水/气温变化/HDD/CDD     ║
║  E. 电网市场   {ce:3d} 列   紫  竞价空间/备用/联络线          ║
╠══════════════════════════════════════════════════════════╣
║  数据源: {source:<44s} ║
║  天气:   {len(NODE_NAMES)} 子节点 → {len(CLUSTERS)} 集群 → {len(VAR_PREFIXES)} 变量  ║
╚══════════════════════════════════════════════════════════╝
""")

    return fp


if __name__ == "__main__":
    export()
